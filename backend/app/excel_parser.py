import os
import re
import json
from decimal import Decimal
from sqlalchemy.orm import Session
import openpyxl
import xlwt
from app.config import settings
from app.models import Company, Employee, SalaryRecord, DeclareTask, ImportBatch, RiskAlert

def parse_group_excel(file_path: str, month: str, db: Session, batch_id: int):
    # Determine the sheet name based on the month (e.g. "2026-07" -> "7月")
    try:
        parts = month.split("-")
        month_num = str(int(parts[1]))  # "07" -> "7"
        target_sheet = f"{month_num}月"
    except Exception:
        target_sheet = "7月"  # Fallback

    wb = openpyxl.load_workbook(file_path, data_only=True)
    if target_sheet not in wb.sheetnames:
        # Try lowercase or search
        found = False
        for s in wb.sheetnames:
            if month_num in s:
                target_sheet = s
                found = True
                break
        if not found:
            raise ValueError(f"Excel文件中未找到与月份 '{month}' 对应的Sheet (如 {target_sheet})。可用Sheet: {wb.sheetnames}")

    ws = wb[target_sheet]
    
    # Load headers (read without values_only=True to allow styling checks later)
    header_row = None
    for row in ws.iter_rows():
        row_values = [x.value for x in row]
        if row_values and any(x is not None for x in row_values):
            header_row = [str(x).strip() if x is not None else "" for x in row_values]
            break
            
    if not header_row:
        raise ValueError("Excel Sheet中未找到表头。")

    # Define standard column name mappings (flexible search)
    col_idx = {
        "company_name": -1,
        "name": -1,
        "id_card": -1,
        "salary": -1,
        "tax_amount": -1,
        "pension": -1,
        "medical": -1,
        "unemployment": -1,
        "actual_tax": -1,
        "employee_status": -1,
        "phone": -1,
        "store_name": -1,
        "post_name": -1,
        "remark": -1
    }

    # Find columns by checking aliases
    for i, col in enumerate(header_row):
        col_lower = col.lower()
        if "公司" in col_lower or "分公司" in col_lower:
            col_idx["company_name"] = i
        elif "姓名" in col_lower and "门店+姓名" not in col_lower:
            col_idx["name"] = i
        elif "身份证" in col_lower or "证件号码" in col_lower:
            col_idx["id_card"] = i
        elif "报税工资" in col_lower or "工资" in col_lower or "应发" in col_lower or "本期收入" in col_lower:
            if col_idx["salary"] == -1 or "报税" in col_lower:
                col_idx["salary"] = i
        elif "个税" in col_lower and "实际个税" not in col_lower:
            col_idx["tax_amount"] = i
        elif "养老" in col_lower:
            col_idx["pension"] = i
        elif "医疗" in col_lower:
            col_idx["medical"] = i
        elif "失业" in col_lower:
            col_idx["unemployment"] = i
        elif "实际个税" in col_lower:
            col_idx["actual_tax"] = i
        elif "状态" in col_lower or "员工状态" in col_lower:
            col_idx["employee_status"] = i
        elif "联系方式" in col_lower or "电话" in col_lower or "手机" in col_lower:
            col_idx["phone"] = i
        elif "门店" in col_lower:
            col_idx["store_name"] = i
        elif "岗位" in col_lower or "职位" in col_lower:
            col_idx["post_name"] = i
        elif "备注" in col_lower:
            col_idx["remark"] = i

    # Validate essential columns
    missing = []
    if col_idx["name"] == -1: missing.append("姓名")
    if col_idx["id_card"] == -1: missing.append("身份证号")
    if col_idx["company_name"] == -1: missing.append("分公司/公司名称")
    if missing:
        raise ValueError(f"Excel表头中缺少必需的列: {', '.join(missing)}。检测到的表头为: {header_row}")

    from sqlalchemy import text
    # 0. Clean existing data for this month to ensure idempotency and prevent duplicate records
    # Use raw SQL to bypass ORM session latency/ghost writes.
    # Note: We DO NOT delete older ImportBatch records so that the import history logs remain in the frontend.
    db.execute(text("DELETE FROM tax_salary_record WHERE month = :m"), {"m": month})
    db.execute(text("DELETE FROM tax_declare_task WHERE month = :m"), {"m": month})
    db.execute(text("DELETE FROM tax_risk_alert WHERE month = :m"), {"m": month})
    
    # Force commit to physically wipe out old data from the DB before writing new rows
    db.commit()
    
    # CRITICAL: Close session to completely discard the ORM Identity Map and wipe out all relationship caches/object graphs.
    # The session will automatically reopen when we query next.
    db.close()

    # 0. Load companies and employees into memory cache to avoid 900+ SELECT queries
    companies_cache = {c.company_name: c for c in db.query(Company).all()}
    employees_cache = {(emp.id_card, emp.company_id): emp for emp in db.query(Employee).all()}

    # To calculate declaration totals per company
    company_totals = {}
    ignored_list = []

    rows_processed = 0
    valid_rows_count = 0

    # Start parsing data rows
    passed_header = False
    for row in ws.iter_rows():
        row_values = [x.value for x in row]
        row_str = [str(x).strip() if x is not None else "" for x in row_values]
        if not passed_header:
            if row_str == header_row:
                passed_header = True
            continue

        if not any(x is not None for x in row_values):
            continue

        # Extract values
        c_name = str(row[col_idx["company_name"]].value).strip() if row[col_idx["company_name"]].value is not None else ""
        emp_name = str(row[col_idx["name"]].value).strip() if row[col_idx["name"]].value is not None else ""
        id_card = str(row[col_idx["id_card"]].value).strip() if row[col_idx["id_card"]].value is not None else ""

        # Normalize ID card
        id_card = id_card.replace(" ", "").upper()

        if not c_name or not emp_name or not id_card:
            continue

        rows_processed += 1

        # Check if cell has RED background color
        name_cell = row[col_idx["name"]]
        is_red = False
        if name_cell.fill and name_cell.fill.fill_type:
            if hasattr(name_cell.fill, 'start_color') and name_cell.fill.start_color:
                if name_cell.fill.start_color.rgb in ["FFFF0000", "FF0000"]:
                    is_red = True

        # Helper to convert to decimal safely
        def to_dec(val):
            if val is None or str(val).strip() == "" or str(val).strip() == "None":
                return Decimal("0.00")
            try:
                clean_val = re.sub(r'[^\d\.\-]', '', str(val))
                return Decimal(clean_val)
            except Exception:
                return Decimal("0.00")

        salary = to_dec(row[col_idx["salary"]].value) if col_idx["salary"] != -1 else Decimal("0.00")
        tax_amount = to_dec(row[col_idx["tax_amount"]].value) if col_idx["tax_amount"] != -1 else Decimal("0.00")
        pension = to_dec(row[col_idx["pension"]].value) if col_idx["pension"] != -1 else Decimal("0.00")
        medical = to_dec(row[col_idx["medical"]].value) if col_idx["medical"] != -1 else Decimal("0.00")
        unemployment = to_dec(row[col_idx["unemployment"]].value) if col_idx["unemployment"] != -1 else Decimal("0.00")
        actual_tax = to_dec(row[col_idx["actual_tax"]].value) if col_idx["actual_tax"] != -1 else Decimal("0.00")
        
        # If actual_tax is 0 and tax_amount is not, use tax_amount
        if actual_tax == Decimal("0.00") and tax_amount != Decimal("0.00"):
            actual_tax = tax_amount

        emp_status = str(row[col_idx["employee_status"]].value).strip() if col_idx["employee_status"] != -1 and row[col_idx["employee_status"]].value is not None else "在职"
        phone = str(row[col_idx["phone"]].value).strip() if col_idx["phone"] != -1 and row[col_idx["phone"]].value is not None else ""
        store_name = str(row[col_idx["store_name"]].value).strip() if col_idx["store_name"] != -1 and row[col_idx["store_name"]].value is not None else ""
        post_name = str(row[col_idx["post_name"]].value).strip() if col_idx["post_name"] != -1 and row[col_idx["post_name"]].value is not None else ""
        remark = str(row[col_idx["remark"]].value).strip() if col_idx["remark"] != -1 and row[col_idx["remark"]].value is not None else ""

        # If red background, skip import but log to ignored_list
        if is_red:
            ignored_list.append({
                "name": emp_name,
                "id_card": id_card,
                "company_name": c_name,
                "salary": float(salary),
                "tax_amount": float(actual_tax),
                "store_name": store_name,
                "post_name": post_name
            })
            continue

        # 1. Check & Auto-create Company (using cache)
        company = companies_cache.get(c_name)
        if not company:
            company = Company(
                company_name=c_name,
                company_code=f"COMP_{c_name[:10]}",
                status="ACTIVE"
            )
            db.add(company)
            db.flush()
            companies_cache[c_name] = company

        # 2. Check & Auto-create / Update Employee (using cache)
        employee = employees_cache.get((id_card, company.id))
        if not employee:
            employee = Employee(
                company_id=company.id,
                name=emp_name,
                id_card=id_card,
                employee_status=emp_status,
                phone=phone
            )
            db.add(employee)
            db.flush()
            employees_cache[(id_card, company.id)] = employee
        else:
            # Update info if changed
            has_changed = False
            if employee.name != emp_name:
                employee.name = emp_name
                has_changed = True
            if employee.employee_status != emp_status:
                employee.employee_status = emp_status
                has_changed = True
            if phone and employee.phone != phone:
                employee.phone = phone
                has_changed = True
            if has_changed:
                db.flush()

        # 3. Create Salary Record
        salary_rec = SalaryRecord(
            batch_id=batch_id,
            company_id=company.id,
            employee_id=employee.id,
            name=emp_name,
            id_card=id_card,
            store_name=store_name,
            post_name=post_name,
            employee_status=emp_status,
            salary=salary,
            tax_amount=tax_amount,
            pension=pension,
            medical=medical,
            unemployment=unemployment,
            actual_tax=actual_tax,
            phone=phone,
            remark=remark,
            month=month
        )
        db.add(salary_rec)
        valid_rows_count += 1

        # Aggregate declaration total
        if company.id not in company_totals:
            company_totals[company.id] = {"tax": Decimal("0.00"), "headcount": 0}
        
        company_totals[company.id]["tax"] += actual_tax
        company_totals[company.id]["headcount"] += 1

    # 4. Generate/Update declaration tasks for each company in the batch
    for comp_id, totals in company_totals.items():
        task = db.query(DeclareTask).filter(
            DeclareTask.month == month,
            DeclareTask.company_id == comp_id
        ).first()
        
        if not task:
            task = DeclareTask(
                month=month,
                company_id=comp_id,
                status="DATA_READY",
                declare_tax=totals["tax"],
                headcount=totals["headcount"]
            )
            db.add(task)
        else:
            # Update existing task
            task.declare_tax = totals["tax"]
            task.headcount = totals["headcount"]
            if task.status == "INIT":
                task.status = "DATA_READY"
        db.flush()

    # Update Batch info
    batch = db.query(ImportBatch).filter(ImportBatch.id == batch_id).first()
    if batch:
        batch.total_rows = rows_processed
        batch.valid_rows = valid_rows_count
        batch.ignored_records = json.dumps(ignored_list, ensure_ascii=False) if ignored_list else None
        batch.status = "SUCCESS"

    db.commit()
    return rows_processed, valid_rows_count

def export_tax_client_excel(company_id: int, month: str, db: Session) -> str:
    # 1. Fetch records
    records = db.query(SalaryRecord).filter(
        SalaryRecord.company_id == company_id,
        SalaryRecord.month == month
    ).all()
    
    company = db.query(Company).filter(Company.id == company_id).first()
    c_name = company.company_name if company else "未知公司"

    # Create XLS workbook and sheet
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('正常工资薪金收入')

    # Columns based on 正常工资薪金所得(1).xls template
    columns = [
        '工号', '*姓名', '*证件类型', '*证件号码', '本期收入', '本期免税收入', 
        '基本养老保险费', '基本医疗保险费', '失业保险费', '住房公积金', 
        '累计子女教育', '累计继续教育', '累计住房贷款利息', '累计住房租金', '累计赡养老人', 
        '累计3岁以下婴幼儿照护', '累计个人养老金', '企业(职业)年金', '商业健康保险', 
        '税延养老保险', '公务交通费用', '通讯费用', '律师办案费用', 
        '西藏附加减除费用', '其他', '准予扣除的捐赠额', '减免税额', '备注'
    ]

    # Set styles
    font_bold = xlwt.Font()
    font_bold.bold = True
    style_header = xlwt.XFStyle()
    style_header.font = font_bold

    # Write header
    for col_num, col_title in enumerate(columns):
        ws.write(0, col_num, col_title, style_header)

    # Write records
    for row_num, rec in enumerate(records, start=1):
        ws.write(row_num, 0, "")                     # 工号
        ws.write(row_num, 1, rec.name)               # *姓名
        ws.write(row_num, 2, "居民身份证")            # *证件类型 (default)
        ws.write(row_num, 3, rec.id_card)            # *证件号码
        ws.write(row_num, 4, float(rec.salary))      # 本期收入
        ws.write(row_num, 5, "")                     # 本期免税收入
        
        # Social security (only write if > 0)
        pension = float(rec.pension) if rec.pension else 0.0
        medical = float(rec.medical) if rec.medical else 0.0
        unemployment = float(rec.unemployment) if rec.unemployment else 0.0
        
        ws.write(row_num, 6, pension if pension > 0 else "") # 基本养老保险费
        ws.write(row_num, 7, medical if medical > 0 else "") # 基本医疗保险费
        ws.write(row_num, 8, unemployment if unemployment > 0 else "") # 失业保险费
        
        ws.write(row_num, 9, "")                     # 住房公积金
        
        # Fill standard empty cells for deductions
        for col_idx in range(10, 27):
            ws.write(row_num, col_idx, "")
            
        ws.write(row_num, 27, rec.remark or "")      # 备注

    # Ensure output folder exists
    temp_dir = os.path.join(settings.UPLOAD_DIR, "exports", month)
    os.makedirs(temp_dir, exist_ok=True)
    
    safe_cname = re.sub(r'[^\w\u4e00-\u9fa5\-]', '_', c_name)
    file_name = f"{safe_cname}_{month}_正常工资薪金所得.xls"
    dest_path = os.path.join(temp_dir, file_name)
    
    wb.save(dest_path)
    return dest_path
